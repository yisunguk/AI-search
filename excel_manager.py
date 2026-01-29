import streamlit as st
import pandas as pd
import openpyxl
from io import BytesIO
import base64
import glob
import os
import os
import json
import fitz # PyMuPDF for DRM check
import zipfile # For Office DRM check

# --- Default Configuration ---
DEFAULT_MAPPING_RULES = {
    "Service of Unit": ["I8"],
    "Item No.": ["AV8"],
    "Size": [ ["E9", "M9", "N9"] ], 
    "Type": ["Y9", "V49"], 
    "Surf/Unit (Gross/Eff)": [ ["K10", "O10", "P10"] ],
    "Fluid Name": ["T13", "AR13"], 
    "Fluid Quantity, Total": ["T14", "AR14"],
    "Temperature (In/Out)": [ 
        {"action": "vertical", "cells": ["T20", "AF20"]}, 
        {"action": "vertical", "cells": ["AR20", "BD20"]} 
    ],
    "Inlet Pressure": ["AB28", "AZ28"],
    "Velocity": ["AB29", "AZ29"],
    "Pressure Drop, Allow/Calc": [ 
        {"action": "vertical", "cells": ["T30", "AF30"]}, 
        {"action": "vertical", "cells": ["AR30", "BD30"]} 
    ],
    "Heat Exchanged": ["M32"],
    "MTD (Corrected)": ["BB32"],
    "Transfer Rate, Service": ["M33"],
    "Clean": ["AH33"],
    "Actual": ["BB33"],
    "Design/Test Pressure": ["T36"],
    "Design Temperature": ["T37"],
    "No Passes per Shell": ["T38"],
    "Tube No.": ["F43"],
    "OD": ["N43", "AC45"], 
    "Thk(Avg)": ["AC43"],
    "Length": ["AR43"],
    "Pitch": ["BG43"],
    "Tube Type": ["F44"],
    "Material": ["AH44"],
    "Tube pattern": ["BM44"],
    "Shell": ["E45"],
    "ID": ["U45"],
    "Shell Cover": ["AU45"],
    "Channel or Bonnet": ["K46"],
    "Channel Cover": ["AU46"],
    "Tubesheet-Stationary": ["K47"],
    "Tubesheet-Floating": ["AW47"],
    "Floating Head Cover": ["K48"],
    "Impingement Plate": ["AW48"],
    "Baffles-Cross": ["H49"],
    "%Cut (Diam)": ["AM49"],
    "Spacing(c/c)": ["AX49"],
    "Inlet": ["BG49"],
    "TEMA Class": ["BA57"]
}

# --- DRM Check Function (Duplicated from app.py to avoid circular imports or refactoring complexity) ---
def is_drm_protected(uploaded_file):
    """
    Check if the uploaded file is DRM protected or encrypted.
    Returns True if protected, False otherwise.
    """
    try:
        file_type = uploaded_file.name.split('.')[-1].lower()
        
        # 1. PDF Check
        if file_type == 'pdf':
            try:
                # Read file stream
                bytes_data = uploaded_file.getvalue()
                with fitz.open(stream=bytes_data, filetype="pdf") as doc:
                    if doc.is_encrypted:
                        return True
            except Exception as e:
                print(f"PDF DRM Check Error: {e}")
                # If we can't open it with fitz, it might be corrupted or heavily encrypted
                return False 

        # 2. Office Files (docx, pptx, xlsx) Check
        elif file_type in ['docx', 'pptx', 'xlsx']:
            try:
                bytes_data = uploaded_file.getvalue()
                # Check if it is a valid zip file
                if not zipfile.is_zipfile(BytesIO(bytes_data)):
                    # Not a zip -> Likely Encrypted/DRM (OLE format)
                    return True
                
                # Optional: Try to open it to be sure
                with zipfile.ZipFile(BytesIO(bytes_data)) as zf:
                    # Check for standard OOXML structure (e.g., [Content_Types].xml)
                    if '[Content_Types].xml' not in zf.namelist():
                        return True
            except Exception as e:
                print(f"Office DRM Check Error: {e}")
                return True # Assume protected if we can't parse structure
                
        return False
    except Exception as e:
        print(f"General DRM Check Error: {e}")
        return False

# --- Helper Functions ---

def rules_to_df(rules_dict):
    """Convert mapping rules dict to a flat DataFrame for editing."""
    rows = []
    for label, rule_list in rules_dict.items():
        for idx, rule in enumerate(rule_list):
            row = {
                "Label": label,
                "Order": idx + 1,
                "Type": "Single",
                "Cells": ""
            }
            
            if isinstance(rule, list):
                row["Type"] = "Merge"
                row["Cells"] = ", ".join(rule)
            elif isinstance(rule, dict) and rule.get("action") == "vertical":
                row["Type"] = "Vertical"
                row["Cells"] = ", ".join(rule["cells"])
            else:
                row["Type"] = "Single"
                row["Cells"] = str(rule)
            
            rows.append(row)
    
    return pd.DataFrame(rows)

def df_to_rules(df):
    """Convert edited DataFrame back to mapping rules dict."""
    rules_dict = {}
    
    # Sort by Label and Order to ensure correct list order
    df = df.sort_values(by=["Label", "Order"])
    
    for _, row in df.iterrows():
        label = row["Label"]
        rtype = row["Type"]
        cells_str = row["Cells"]
        
        # Parse cells
        cells = [c.strip() for c in cells_str.split(",") if c.strip()]
        
        if not cells:
            continue
            
        rule = None
        if rtype == "Merge":
            rule = cells
        elif rtype == "Vertical":
            rule = {"action": "vertical", "cells": cells}
        else: # Single
            rule = cells[0]
            
        if label not in rules_dict:
            rules_dict[label] = []
        
        rules_dict[label].append(rule)
        
    return rules_dict

def find_template_file(base_dir="Excel"):
    """Find template file in the specified directory."""
    # Check if directory exists
    if not os.path.exists(base_dir):
        return None
        
    # Look for template.xlsx in the Excel directory
    target_path = os.path.join(base_dir, "template.xlsx")
    if os.path.exists(target_path): 
        return target_path
        
    for file in os.listdir(base_dir):
        if file.lower() == "template.xlsx": 
            return os.path.join(base_dir, file)
            
    xlsx_files = [f for f in os.listdir(base_dir) if f.endswith(".xlsx")]
    ignore_list = ["processed_output.xlsx", "dummy_input.xlsx", "dummy_template.xlsx"]
    candidates = [f for f in xlsx_files if f not in ignore_list and not f.startswith("dummy_") and not f.startswith("~$")]
    return os.path.join(base_dir, candidates[0]) if candidates else None

def get_cell_value(sheet, addr):
    val = sheet[addr].value
    return str(val) if val is not None else ""

def process_excel(input_file, template_file, mapping_rules):
    template_wb = openpyxl.load_workbook(template_file)
    template_sheet = template_wb.active
    
    input_wb = openpyxl.load_workbook(input_file, data_only=True)
    input_sheet_names = input_wb.sheetnames
    
    for i, sheet_name in enumerate(input_sheet_names):
        input_sheet = input_wb[sheet_name]
        target_col_idx = 3 + i
        template_sheet.cell(row=1, column=target_col_idx).value = i + 1
        
        duplicate_counters = {key: 0 for key in mapping_rules}
        
        for row_idx in range(2, 150):
            label_cell = template_sheet.cell(row=row_idx, column=1)
            label = label_cell.value
            
            if label:
                label = str(label).strip()
                if label in mapping_rules:
                    rules = mapping_rules[label]
                    counter = duplicate_counters[label]
                    
                    if counter < len(rules):
                        rule = rules[counter]
                        try:
                            if isinstance(rule, dict) and rule.get("action") == "vertical":
                                cells = rule["cells"]
                                val1 = get_cell_value(input_sheet, cells[0])
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = val1
                                if len(cells) > 1:
                                    val2 = get_cell_value(input_sheet, cells[1])
                                    template_sheet.cell(row=row_idx + 1, column=target_col_idx).value = val2
                            elif isinstance(rule, list):
                                values = [get_cell_value(input_sheet, addr) for addr in rule]
                                merged_value = " ".join([v for v in values if v])
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = merged_value
                            else:
                                val = get_cell_value(input_sheet, rule)
                                template_sheet.cell(row=row_idx, column=target_col_idx).value = val
                        except Exception as e:
                            print(f"Error processing {label}: {e}")
                        duplicate_counters[label] += 1

    output = BytesIO()
    template_wb.save(output)
    output.seek(0)
    return output

def render_excel_tool():
    st.title("엑셀 데이터 자동 추출 (Excel Auto-Filler)")
    
    # Initialize Session State
    if "mapping_rules" not in st.session_state:
        st.session_state.mapping_rules = DEFAULT_MAPPING_RULES
        
    # --- Persistent Error Display ---
    if "excel_drm_error" in st.session_state and st.session_state.excel_drm_error:
        st.error(st.session_state.excel_drm_error)
        del st.session_state.excel_drm_error

    # Tabs
    tab1, tab2 = st.tabs(["📂 데이터 처리 (Data Processing)", "⚙️ 매핑 설정 (Mapping Settings)"])

    # --- Tab 1: Processing ---
    with tab1:
        st.markdown("### 파일 업로드 및 처리")
        
        col1, col2 = st.columns(2)
        with col1:
            if "excel_input_key" not in st.session_state:
                st.session_state.excel_input_key = 0
            input_file = st.file_uploader("1. 입력 파일 업로드 (Data)", type=['xlsx'], key=f"excel_input_{st.session_state.excel_input_key}")
        with col2:
            # Look for template in Excel folder
            found_template = find_template_file()
            if found_template:
                template_file = found_template
                st.info(f"기본 템플릿 사용 중: '{os.path.basename(found_template)}'")
            else:
                if "excel_template_key" not in st.session_state:
                    st.session_state.excel_template_key = 0
                template_file = st.file_uploader("2. 템플릿 파일 업로드 (Form)", type=['xlsx'], key=f"excel_template_{st.session_state.excel_template_key}")
                st.warning("기본 템플릿을 찾을 수 없습니다. 파일을 업로드해주세요.")

        if input_file and template_file:
            if st.button("엑셀 파일 처리 시작 (Process)", type="primary"):
                # DRM Check
                if is_drm_protected(input_file):
                    st.session_state.excel_drm_error = f"⛔ 입력 파일({input_file.name})이 DRM으로 보호되어 있습니다. 파일 목록에서 제거되었습니다."
                    st.session_state.excel_input_key += 1
                    st.rerun()
                elif is_drm_protected(template_file) and not isinstance(template_file, str): # template_file could be path string
                    st.session_state.excel_drm_error = f"⛔ 템플릿 파일({template_file.name})이 DRM으로 보호되어 있습니다. 파일 목록에서 제거되었습니다."
                    if "excel_template_key" in st.session_state:
                        st.session_state.excel_template_key += 1
                    st.rerun()
                else:
                    try:
                        with st.spinner("처리 중..."):
                            result_file = process_excel(input_file, template_file, st.session_state.mapping_rules)
                        st.success("처리 완료!")
                        st.download_button(
                            label="결과 엑셀 다운로드",
                            data=result_file,
                            file_name="processed_output.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    except Exception as e:
                        st.error(f"오류 발생: {e}")

    # --- Tab 2: Settings ---
    with tab2:
        st.markdown("### ⚙️ 매핑 규칙 설정")
        st.markdown("""
        - **Label**: 템플릿 파일의 A열에 있는 텍스트입니다.
        - **Type**: 
            - `Single`: 단일 셀 (예: I8).
            - `Merge`: 여러 셀 병합 (예: E9, M9).
            - `Vertical`: 두 행으로 분할 (예: In/Out Temp).
        - **Cells**: 쉼표로 구분된 셀 주소.
        - **Order**: 라벨이 여러 번 나올 경우 순서 (1, 2, 3...).
        """)
        
        # Convert current rules to DF
        current_df = rules_to_df(st.session_state.mapping_rules)
        
        # Data Editor
        edited_df = st.data_editor(
            current_df,
            num_rows="dynamic",
            column_config={
                "Type": st.column_config.SelectboxColumn(
                    "Mapping Type",
                    options=["Single", "Merge", "Vertical"],
                    required=True
                ),
                "Order": st.column_config.NumberColumn(
                    "Order",
                    min_value=1,
                    step=1,
                    required=True
                )
            },
            use_container_width=True,
            hide_index=True
        )
        
        # Save Button
        if st.button("설정 저장 (현재 세션에 적용)"):
            try:
                new_rules = df_to_rules(edited_df)
                st.session_state.mapping_rules = new_rules
                st.success("설정이 적용되었습니다! '데이터 처리' 탭에서 사용할 수 있습니다.")
            except Exception as e:
                st.error(f"설정 저장 실패: {e}")
                
        st.markdown("---")
        st.markdown("### 💾 설정 백업 및 복원")
        
        # Export
        current_rules_json = json.dumps(st.session_state.mapping_rules, indent=4, ensure_ascii=False)
        st.download_button(
            label="설정 다운로드 (JSON)",
            data=current_rules_json,
            file_name="mapping_config.json",
            mime="application/json"
        )
        
        # Import
        uploaded_config = st.file_uploader("설정 불러오기 (JSON)", type=['json'])
        if uploaded_config:
            try:
                loaded_rules = json.load(uploaded_config)
                if st.button("불러온 설정 적용"):
                    st.session_state.mapping_rules = loaded_rules
                    st.success("설정이 성공적으로 로드되었습니다!")
                    st.rerun()
            except Exception as e:
                st.error(f"JSON 로드 실패: {e}")
